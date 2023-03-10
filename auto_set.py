# 导入模块
import openpyxl
import re
import pandas as pd
import json
import os
import chardet
import csv

# 定义json_to_excel函数 此处处理为：json内容位于每行的最后，将其向右展开并添加json对应键的表头 按需修改
def json_to_excel(wb):
    ws = wb.active

    # 获取最大行数和列数
    max_row = ws.max_row
    max_col = ws.max_column
    # 定义正则表达式匹配json字符串
    pattern = re.compile(r"\{[\s\S]*?\}")

    # 遍历最后一列的单元格，提取json字符串，并生成新表格
    for i in range(2, max_row + 1):
        # 跳过标题行
        cell = ws.cell(row=i, column=max_col)
        if cell.row == 1:
            continue
        # 获取单元格内容
        value = cell.value
        if value == None:
            continue
        # 匹配json字符串
        match = pattern.search(value)
        # 如果匹配成功，提取json字符串，并转换为字典
        if match:
            json_str = match.group()
            json_str = json_str.replace('_x000D_','\r')
            json_dict = eval(json_str)

            # 根据字典生成新表格
            new_table = pd.DataFrame({key: [value] for key, value in json_dict.items()})

            cell_dict = json.loads(json_str)
            keys = list(cell_dict.keys())
            values = list(cell_dict.values())
            # 在第一行末尾添加键列表中的元素作为新的列名，并在相应的单元格下方添加值列表中的元素作为新数据
            for j in range(len(keys)):
                # 计算新列名所在位置
                new_col_index = max_col + j
                # 在第一行末尾添加新列名
                ws.cell(row=1, column=new_col_index).value = keys[j]
                # 在相应位置添加新数据
                ws.cell(row=i, column=new_col_index).value = values[j]
    return wb

# CSV转换Excel
csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
for file in csv_files:
    encoding = chardet.detect(open(file, 'rb').read())['encoding']
    print("文件编码：",encoding)
    
    #编码范围扩展
    if(encoding=='GB2312'):
        encoding='GB18030'
    #
    #   针对以=不同分隔符的读取
    try:
        df = pd.read_csv(file, encoding=encoding, sep=',')
    except:
        try:
            df = pd.read_csv(file, encoding=encoding, sep='\t')
        except:
            print("无合法分隔符")

    xlsx_name = file.replace('.csv', '.xlsx')
    df.to_excel(xlsx_name, index=False)

# 获取当前目录下的所有xlsx文件
files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
print("找到的xlsx文件：",files)

# 遍历每个文件
for file in files:
    # 打印提示信息
    print(f'正在处理{file}...')
    
    try:
        # 打开文件
        wb = openpyxl.load_workbook(file)
        # 调用json_to_excel函数
        wb = json_to_excel(wb)
        # 保存修改后的文件
        new_file = file[:-5] + '-统计表.xlsx'
        wb.save(new_file)
        # 打印提示信息
        print(f'{file}处理完成，转换为{new_file}')
        # 删除原xlsx
        os.remove(file)
    except:
        print(f'处理{file}失败,尝试更改文件编码或转换为xlsx文件')

#暂停
input('按任意键退出')
